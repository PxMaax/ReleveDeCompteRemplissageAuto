import openpyxl as pyxl
import glob
import tkinter as tk
import tkinter.font
from tkinter import font
import csv
import os
from tkinter import filedialog
import re
from datetime import date
from erreur_class import ErreurExcel
from datetime import datetime


bg_color = "#55868C"
button_color = "#7F636E"
text_color = "#000000"

largeur_fenetre = "600"
hauteur_fenetre = "500"
fenetreResolution = largeur_fenetre + "x" + hauteur_fenetre

policyHeader="Baskerville"
sizeHeader=20
weightHeader= "bold"

policyBody="Baskerville"
sizeBody=14
weightBody= "bold"

# Obtenir la date d'aujourd'hui
aujourd_hui = date.today()

# Convertir la date en chaîne de caractères
date_str = aujourd_hui.strftime("%d/%m/%Y")  # Format: jour/mois/année

# Espacement
y_spacing = 40  # Espacement vertical entre les éléments
x_spacing = -10 # Espacement horizontal entre les élements

class ReleveAuto:
    
    mois = ""
    annee = ""
    lignedebut = 0
    version = "Version 1"
    nbJours = 0
    jour = 0
        
    fileFeuilleDeCompta = ""
    fileReleveDeCompte = ""
    
    colone_depart_releve = "B"
    ligne_releve = 11
    
    tableau_de_compta_wb = None
    releve_de_compte_wb  = None
    
    tableau_de_compta_ws = None
    releve_de_compte_ws = None
    
    excel_des_problemes = pyxl.Workbook()
    sheet_des_problemes = excel_des_problemes.active
    sheet_des_problemes['A1'] = "Case"
    sheet_des_problemes['B1'] = "Erreur"
    
    tableau_erreur = []
    
    case_iteration = ""
    numMois = 0
    
    
    def fill_fichier_error (self):
        for index, (valeur_1, valeur_2) in enumerate(self.tableau_erreur, start=2):
            self.sheet_des_problemes.cell(row=index, column=1, value=valeur_1)
            self.sheet_des_problemes.cell(row=index, column=2, value=valeur_2)
        
    def noter_releve_de_compte(self,ligne,message):
        self.releve_de_compte_ws["E" + ligne] = message
        return 
    
    def verif_CB(self,valeur_SC,valeur_AvC,day_value):
        
        """
            verif_CB : Check les valeurs trouvées dans le relevé de compte match avec les fiches de caisse

            :param valeur_SC : valeur carte Sans contact
            :param valeur_AvC : valeur carte Avc Contact
            :return : Vrai si match
         """ 
        print("valeur sc" +  str(valeur_SC))
        print("valeur AvC" +  str(valeur_AvC))
        print("jour : " + str(day_value))
        print("valeur dans le tableau de compte à la case : " + "D"+str(self.lignedebut+day_value-1) + " = " + str(self.tableau_de_compta_ws["D"+str(self.lignedebut+day_value-1)].value))
        print("la valeur entre les deux cartes bleues" + str(round((valeur_SC + valeur_AvC),2)))
        
        if (str(self.tableau_de_compta_ws["D"+str(self.lignedebut+day_value-1)].value) != str(round((valeur_SC + valeur_AvC),2))):
            raise ErreurExcel(self.case_iteration, "La valeur dans le relevé et dans l'excel de compta ne correspondent pas")
        return True
    
    def getValeurCredit (self, current_cell):
        
        """
            getValeurCredit : Function trouve la valeur de crédit
            :param current_cell : case de la valeur qu'on cherche
            :return: Credit lié à la case en cours
        """ 
        current_cell.row
        return self.releve_de_compte_ws["D" +str(current_cell.row)].value
        
    def extract_match_date_from_string(self,value_case_releve):
        # Utiliser une expression régulière pour rechercher la date au format "DD/MM"
        date_pattern = r'\b\d{2}/\d{2}\b'
        match = re.search(date_pattern, value_case_releve)
        date_string = match.group()
        if date_string:
            return date_string
        else:
            raise ErreurExcel(self.case_iteration, "la date n'a pas été trouvée")
 
    def trouver_case_carte_meme_date(self, ligne_current_cell, target_date):
        
        """
            trouver_case_carte_meme_date : Function qui permet de trouver l'autre carte bleue pour compléter la journée
            
            La fonction check dans les +15/-15 cases autour si elle contient le code des cartes sans contact
            
            :param ligne_current_cell : la ligne de la case courrante
            :param target_date : la date à trouver dans les cases
            :return: valeur de la CB SC
        """ 
        
        # Coordonnées de la case actuelle
        flag = False
        # print("target date : " + target_date )
        # print("ligne_current_cell" + str(ligne_current_cell))
        # print( "current row")
        # print(current_row)
        # print("target value")
        # print(target_value)
        
        # # Parcourir les 15 cases au-dessus et en dessous de la case actuelle
        for row_offset in range(-15, 16):
                # print("row offset :" + str(row_offset))
                # Coordonnées de la case à vérifier
                row_to_check = ligne_current_cell + row_offset
                # print('row to check ')
                # print(row_to_check)
                # Récupérer la valeur et la date dans la case à vérifier
                if row_to_check >0:
                    cell_value = self.releve_de_compte_ws.cell(row=row_to_check, column=2).value
                    # print("case : B" + str(row_to_check)  )
                
                    if (cell_value is not None ) and ("0697462" in cell_value) and (target_date in cell_value):
                        cell_date = self.releve_de_compte_ws.cell(row=row_to_check, column=2)
                        # print("cell date")
                        # print(cell_date.value)
                        if flag is True :
                            raise ErreurExcel(self.case_iteration,"Plusieurs cartes trouvées pour la date : " + target_date)
                        else: flag = True
                            
                    
        if flag is False:
            date_object = datetime.strptime(target_date, '%d/%m')
            day_value = date_object.day
            if self.verif_CB(0,self.getValeurCredit(self.releve_de_compte_ws[self.case_iteration]),day_value) == False :
                raise ErreurExcel(self.case_iteration,"Aucune carte avec contact n'a été trouvée à cette date : " + target_date)
            else: return self.releve_de_compte_ws["AAAA10000"]
        return cell_date

    ## lire la case, analyser son contenu, mettre les bonnes valeurs au bon endroit, ou reporter valeur dans un excel
    def lecture_ligne_releve(self,ligne_releve, case_releve):
        try:
            value_case_releve = case_releve.value
            ##Si la case est une carte bleue
            if ("REMISE CARTE") in value_case_releve:
                # print("find remise carte in line : "+ str(ligne_releve))
                ## Si la carte est une AvC
                if ("2232935") in value_case_releve:
                       ## print("find card Avc : "+ str(ligne_releve))
                    ## Rechercher la CB SC de la ^m journée
                    ## Ajouter les deux cartes bleues ensemble
                    ## check avec le tableur compta s'il y a match 
                    ## Case cochée si oui, Erreur si non
                    
                    ## match group : format "DD/MM"
                    
                        complete_date = self.extract_match_date_from_string(value_case_releve)

                        if complete_date:
                            # Conversion de la chaîne de date en objet de date Python
                            date_object = datetime.strptime(complete_date, '%d/%m')
                            # Obtenir le jour en tant qu'entier
                            day_value = date_object.day
                            month_value = str(date_object.month)
                            # print('est ce que la valeur month :'+  month_value + 'est egale à ' + self.numMois)
                        if complete_date != "no match" and month_value == self.numMois:
                            case_carte_sc = self.trouver_case_carte_meme_date(ligne_releve,complete_date)                         
                            if self.verif_CB(self.getValeurCredit(case_carte_sc),self.getValeurCredit(self.releve_de_compte_ws[self.case_iteration]),day_value) == True :
                                self.noter_releve_de_compte(str(ligne_releve ), "Vérifiée CB")
                ## DABoi
                elif ("2903075") in value_case_releve:
                    if self.getValeurCredit(self.releve_de_compte_ws[self.case_iteration]) == True :
                        self.noter_releve_de_compte(str(ligne_releve), "Vérifiée DAB")
                    else : 
                        self.noter_releve_de_compte(str(ligne_releve), "Non écrite DAB")
                
        except ErreurExcel as e:  
                self.tableau_erreur.append((e.current_cellCoord, e.details_error))        
                
                return

    def Execution(
        self
    ):
        
        try :
            self.tableau_de_compta_wb = pyxl.load_workbook(self.fileFeuilleDeCompta)
            self.releve_de_compte_wb  = pyxl.load_workbook(self.fileReleveDeCompte)

            for sheet in self.tableau_de_compta_wb:  ## recherche de la bonne feuille de la bonne année
                if sheet.title == str(self.annee):  ##si le nom de la feuille corrsepond à l'année
                    # print("le tableau de compta a été trouvé")
                    self.tableau_de_compta_ws = sheet  ##stockage de la sheet
            
            ## pour stocker et creer la case de départ
            ligne = 1  
            ## pour faire les cases
            colone = "A"  
            flag = True
            
            ## recherche du mois dans le tbleur de compta
            while ligne < 467 and flag == True:  
                casedate = colone + str(ligne)
                # print("self.tableau_de_compta_ws[casedate].value" + self.tableau_de_compta_ws[casedate].value)
                if self.tableau_de_compta_ws[casedate].value == self.mois:
                    # print("le mois a été trouvé")
                    self.lignedebut = (
                        ligne + 3
                    )
                    flag = False
                    colone = "V"
                ligne = ligne + 1
                
                if self.mois == "JANVIER":
                    self.numMois= "1"
                elif self.mois == "FEVRIER":
                    self.numMois= "2"
                elif self.mois == "MARS":
                    self.numMois= "3"
                elif self.mois == "AVRIL":
                    self.numMois= "4"
                elif self.mois == "MAI":
                    self.numMois= "5"
                elif self.mois == "JUIN":
                    self.numMois= "6"
                elif self.mois == "JUILLET":
                    self.numMois= "7"
                elif self.mois == "AOUT":
                    self.numMois= "8"
                elif self.mois == "SEPTEMBRE":
                    self.numMois= "9"
                elif self.mois == "OCTOBRE":
                    self.numMois= "10"
                elif self.mois == "NOVEMBRE":
                    self.numMois= "11"
                elif self.mois == "DECEMBRE":
                    self.numMois= "12"

            ## lire la colone de relevé de compte
            ##case de départ du relevé de compte

            
            self.releve_de_compte_ws = self.releve_de_compte_wb.active
            
            self.case_iteration =  str(self.colone_depart_releve) + str(self.ligne_releve)
            while (self.releve_de_compte_ws[self.case_iteration].value is not None ):
                
                self.lecture_ligne_releve(self.ligne_releve,self.releve_de_compte_ws[self.case_iteration])
                self.ligne_releve = self.ligne_releve + 1
                self.case_iteration = "B" + str(self.ligne_releve)
            
            print("chaussure :) ")  ## j'ai retrouvé mes chaussures!!
            
            fichier_enregistrement = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Fichiers Excel", "*.xlsx"), ("Tous les fichiers", "*.*")]
        )
            
            
            self.fill_fichier_error()
            self.tableau_de_compta_wb.save(self.fileFeuilleDeCompta)  ##save de la feuille de compta
            self.excel_des_problemes.save(fichier_enregistrement)
            self.releve_de_compte_wb.save(self.fileReleveDeCompte)
        except ErreurExcel as e:  
                self.tableau_erreur.append((e.current_cellCoord, e.details_error))      
            
class Application(tk.Tk):
    
    def __init__(self):
        tk.Tk.__init__(self)
        self.resizable(False, False)
        self.iconbitmap("Logo.ico")
        self.title("Relevé de compte automatique")
        self.configure(bg=bg_color)
        self.ReleveAuto = ReleveAuto()
        self.creer_widgets()
        
        # Récupération de la taille de l'écran
        largeur_ecran = self.winfo_screenwidth()
        hauteur_ecran = self.winfo_screenheight()
        x = (largeur_ecran - int(largeur_fenetre)) // 2
        y = (hauteur_ecran - int(hauteur_fenetre)) // 2
        y_spacing = 40  # Espacement vertical entre les éléments
        x_spacing = -10

        # Configuration de la position et de la taille de la fenêtre
        self.geometry(f"{int(largeur_fenetre)}x{int(hauteur_fenetre)}+{x}+{y}")
 
    ## la fonction qui permet de générer tous les éléments du front
    def creer_widgets(self):
        canvas = tk.Canvas(self, bg=bg_color, width=largeur_fenetre, height=hauteur_fenetre)
        canvas.place(x=0, y=0)
        header_label = tk.Label(
            canvas,
            text="Relevé de compte automatique",
            font=font.Font(family=policyHeader, size=sizeHeader, weight=weightHeader),
            bg=bg_color,
            fg=text_color,
            )
        header_label.place(x=100, y=20)
        
        # Rectangle autour de l'entête
        header_rectangle = tk.Canvas(
            canvas, bg=bg_color, highlightbackground="black", highlightthickness=2
        )
        header_rectangle.place(x=95, y=15, width=430, height=48)
        header_label.lift(aboveThis=header_rectangle)
        
        # Bouton Donner feuille de compta
        self.feuille_btn = tk.Button(
            canvas,
            text="Donner feuille de compta",
            command=lambda : self.donnerFile(canvas,"feuille_btn"),
            font=font.Font(family=policyBody, size=sizeBody, weight=weightBody),
            bg=button_color,
            fg=text_color,
        )
        self.feuille_btn.place(x=50 + x_spacing, y=80 + y_spacing, width=250)
        
        self.releve_btn = tk.Button(
            canvas,
            text="Donner relevé de compte",
            command=lambda: self.donnerFile(canvas,"releve_btn"),
            font=font.Font(family=policyBody, size=sizeBody, weight=weightBody),
            bg=button_color,
            fg=text_color,
        )
        self.releve_btn.place(x=50 + x_spacing, y=130 + y_spacing, width=250)
        
        selection_label = tk.Label(
            canvas,
            text="Saisir le mois et l'année:",
            font=font.Font(family=policyBody, size=sizeBody, weight=weightBody),
            bg=bg_color,
            fg=text_color,
        )
        selection_label.place(x=50 + x_spacing, y=180 + y_spacing)
        
        self.month_var = tk.StringVar(canvas)
        self.month_var.set("Mois")  # Ajoutez cette ligne pour définir "Mois" comme valeur par défaut
        month_dropdown = tk.OptionMenu(
            canvas,
            self.month_var,
            "Mois",  # Ajoutez "Mois" comme première option dans la liste
            "JANVIER",
            "FEVRIER",
            "MARS",
            "AVRIL",
            "MAI",
            "JUIN",
            "JUILLET",
            "AOUT",
            "SEPTEMBRE",
            "OCTOBRE",
            "NOVEMBRE",
            "DECEMBRE",
        )

        month_dropdown.config(
            font=font.Font(family=policyBody, size=sizeBody, weight=weightBody), bg=button_color, fg=text_color, highlightthickness=0
        )

        month_dropdown.place(x=50 + x_spacing, y=220 + y_spacing, width=250)

        # Menu déroulant pour l'année
        self.year_var = tk.StringVar(canvas)
        self.year_var.set("Année")
        year_dropdown = tk.OptionMenu(canvas, self.year_var, "2023", "2024", "2025")
        year_dropdown.config(
            font=font.Font(family=policyBody, size=sizeBody, weight=weightBody),bg=button_color, fg=text_color, highlightthickness=0
        )
        year_dropdown.place(x=320 + x_spacing, y=220 + y_spacing, width=150)


        # Affichage du mois et de l'année sélectionnés
        selected_label = tk.Label(
            canvas,
            text="Mois et année sélectionnés:",
            font=font.Font(family=policyBody, size=sizeBody, weight=weightBody),
            bg=bg_color,
            fg=text_color,
        )
        selected_label.place(x=50 + x_spacing, y=280 + y_spacing)
        
        selected_month_label = tk.Label(
            canvas, textvariable=self.month_var, font=font.Font(family=policyBody, size=sizeBody, weight=weightBody), bg=bg_color, fg=text_color
        )
        selected_month_label.place(x=50, y=320 + y_spacing)

        selected_year_label = tk.Label(
            canvas, textvariable=self.year_var, font=font.Font(family=policyBody, size=sizeBody, weight=weightBody), bg=bg_color, fg=text_color
        )
        selected_year_label.place(x=180, y=320 + y_spacing)
        
        # Bouton Lancer le programme
        launch_btn = tk.Button(
            canvas,
            text="Lancer le programme",
            command=lambda:self.TestVariable(),
            font=font.Font(family=policyBody, size=sizeBody, weight=weightBody),
            bg=button_color,
            fg=text_color,
        )
        launch_btn.place(x=20, y=450, width=250)
        
        # Bouton Fermer la page
        close_btn = tk.Button(
            canvas,
            text="Fermer",
            command=self.destroy,
            font=font.Font(family=policyBody, size=sizeBody, weight=weightBody),
            bg=button_color,
            fg=text_color,
        )
        close_btn.place(x=430, y=450, width=150)
        
    ## foncton permetant d'afficher un message d'erreur    
    def afficher_message(parent, message):
    # Fonction pour gérer l'événement du clic sur le bouton "OK"
        def fermer_fenetre():
            message_fenetre.destroy()

        # Création de la fenêtre
        message_fenetre = tk.Toplevel(parent)
        message_fenetre.title("Message d'erreur")
        message_fenetre.iconbitmap("logo.ico")

        # Création d'un label pour afficher le message
        label_message = tk.Label(message_fenetre, text=message, font=font.Font(family=policyBody, size=sizeBody, weight=weightBody))
        label_message.pack(padx=20, pady=20)

        # Création d'un bouton "OK"
        bouton_ok = tk.Button(
            message_fenetre, text="OK", command=fermer_fenetre, width=10, height=2
        )
        bouton_ok.pack(pady=20)

        # Calcul de la taille de la fenêtre en fonction de la longueur du message
        largeur_message_fenetre = max(350, label_message.winfo_reqwidth() + 40)
        hauteur_message_fenetre = 180

        # Récupération de la taille de la fenêtre parente
        parent.update_idletasks()  # Actualisation des tâches du parent
        parent_width = parent.winfo_width()
        parent_height = parent.winfo_height()

        # Calcul de la position de la fenêtre pour la centrer
        xerrorwindow = parent.winfo_rootx() + (parent_width - largeur_message_fenetre) // 2
        yerrorwindow = parent.winfo_rooty() + (parent_height - hauteur_message_fenetre) // 2

        # Configuration de la position et de la taille de la fenêtre
        message_fenetre.geometry(
            f"{largeur_message_fenetre}x{hauteur_message_fenetre}+{xerrorwindow}+{yerrorwindow}"
        )

        # Lancement de la boucle principale de la fenêtre
        message_fenetre.mainloop()

        # Bouton Soumettre le mois et l'annéereleve_btn
        # submit_btn = tk.Button(
        #     canvas,
        #     text="Soumettre",
        #     command=lambda : self.SoumettreMoisAnnee(),
        #     font=font.Font(family=policyBody, size=sizeBody, weight=weightBody),
        #     bg=button_color,
        #     fg=text_color,
        # )
        # submit_btn.place(x=490 + x_spacing, y=220 + y_spacing, width=100)

    ## correspond à la commande effectuée lors du push des 2 premiers boutons
    def donnerFile(self,Cnvas,nomBouton):
        File = filedialog.askopenfilename(initialdir="Desktop/", title="Feuille de compta")
        # print(File)
        
        if nomBouton == "feuille_btn":
            xcercle = int(largeur_fenetre) - 25
            ycercle = self.feuille_btn.winfo_y() + 18
            ## ajout cercle deuxieme bouton
        elif nomBouton == "releve_btn":
            xcercle = int(largeur_fenetre) - 25
            ycercle = self.releve_btn.winfo_y() + 20
            
        rayon = 20    
        
        if ".xlsx" in File:
            if File != "":
                ##Create signal good format
                    Cnvas.create_oval(
                xcercle - rayon,
                ycercle - rayon,
                xcercle + rayon,
                ycercle + rayon,
                fill="green",
                )
                    if nomBouton == "feuille_btn":
                        
                        
                        print("feuille de compta donnée")
                        self.ReleveAuto.fileFeuilleDeCompta = File
                        
                    elif nomBouton == "releve_btn":
                        print("relevé de compte donné")
                        self.ReleveAuto.fileReleveDeCompte = File
                        
        else:
            Cnvas.create_oval(
                xcercle - rayon,
                ycercle - rayon,
                xcercle + rayon,
                ycercle + rayon,
                fill="red",
            )
            self.afficher_message(
                "ERREUR: Extension fichier non reconnue. \n Tu t'es surement trompée de fichier. \n Si le problème persiste, appelle Maxime.",
            )
        return ""
    
    
    

    def TestVariable(self):
        
        self.ReleveAuto.mois = self.month_var.get()
        self.ReleveAuto.annee = self.year_var.get()
        
        if self.ReleveAuto.fileFeuilleDeCompta == "":
            self.afficher_message("Tu n'as pas donné de tableur de compta")

        elif self.ReleveAuto.fileReleveDeCompte == "":
            self.afficher_message(
                "Tu n'as donné de le relevé de compte"
            )

        elif self.month_var.get() == "Mois":
            self.afficher_message("Tu n'as saisi le mois ")
        

        elif self.year_var.get() == "Année":
            self.afficher_message("Tu n'as saisi l'année ")

        else:        
            
            try: 
                # print("mois :" + self.ReleveAuto.mois)
                # print("année :" + self.ReleveAuto.annee)
                self.ReleveAuto.Execution()
                self.afficher_message("Compta effectuée")
                            
            except ErreurExcel as e:  
                self.afficher_message(self,e.details_error)   
                    
        
    
if __name__ == "__main__":
    app = Application()
    app.mainloop()

print("chaussure :) ")  ##j'ai retrouvé mes chaussures !
